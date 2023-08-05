import openpyxl
from openpyxl.styles import PatternFill  # 导入填充模块
import TLB

filename = r'C:\Users\WX\Desktop\WorkFile\项目\内研项目\防干烧\防干烧测试数据\防干烧样机测试数据.xlsx'
Rd = openpyxl.load_workbook(filename)      # 加载存在的excel文件，与Workbook()的区别在于，load_workbook（）未一个方法，一般是加载已存在的文件，Workbook()是创建一个类对象，用于创建一个空白的excel（）
# print(Rd.get_sheet_names())
# print(Rd.sheetnames)                        # 获取文件内的各个工作表名称
xlsx_sheet_number = len(Rd.sheetnames)
# print(xlsx_sheet_number)
Sheet0_data = Rd[Rd.sheetnames[0]]          # 获取工作表0内的所有数据 创建数组变量的目的是防止后面使用该变量时，编译器判定为创建新变量
Array_Len = [0]                             # 获取该工作表内的最大行
List_Len = [0]                              # 获取该工作表内的最大列

DRY_BURNING_THTRSHOLD_FIRST = 0.0035
DRY_BURNING_THTRSHOLD_OTHER = 0.003

Koushu = [0]        # 抽吸口数
Slope_K = [0]       # 阻值变化斜率
Duration = [0]      # 持续时间
Once_Data = [0]*100  # 单口抽吸数据
Suction_Duration = [0]*5000     # 记录每一口的抽吸时长
Average_Time = [0]  # 平均抽吸时长
Save_Koushu_Time_Statistics = [0]*5   # MCU保存的口数


def xlsx_read():
    once_data_adr = 0
    ram_adr = 0
    ram_koushu = 0
    array_3_adr = 1
    Array_Len[0] = Sheet0_data.max_row + 1      # 获取该工作表内的最大行
    List_Len[0] = Sheet0_data.max_column + 1    # 获取该工作表内的最大列
    Error_flag = 0
    K_write_En = 0
    # print("Array_Len=", Array_Len[0])
    # print("List_Len=", List_Len[0])
    # print("Array_Len=", Array_Len)
    # print("List_Len=", List_Len)
    Sheet0_data.column_dimensions['A'].width = 10  # 设置列宽 测试发现该列宽并非设置A列会作用在该工作表全局内，此时设置其它列可以修改该列的列宽，但是未明确设置的会默认为A列的值
    Sheet0_data.column_dimensions['B'].width = 10
    Sheet0_data.column_dimensions['C'].width = 10
    for ram_num in range(1, Array_Len[0], 1):
        rom_data = Sheet0_data.cell(ram_num, 1).value  # 提取Sheets0内第1列num行的数据  注：该库数据读取从1开始。因此第一列对应列表A
        if rom_data == 250:     # 口数起始帧
            if Sheet0_data.cell(ram_num+2, 1).value != 250 or (Sheet0_data.cell(ram_num-2, 1).value == 249 and Sheet0_data.cell(ram_num+2, 1).value == 250):     # 排除MCU保存的口数为250造成的数据干扰
                Koushu[0] += 1
                Sheet0_data.cell(ram_num, 2).value = Koushu[0]  # 将数据 Koushu[0]写入（ram_num，2）单元格
                if once_data_adr != 0:
                    once_data_adr -= 1
                    if ((Once_Data[once_data_adr] * 256 + Once_Data[once_data_adr-1]) - 1) != ram_koushu and ram_koushu != 0:
                        print("口数存储错误!!!", "口数=", Koushu[0]-1, "记录上一口数", ram_koushu, "当前口数", Once_Data[once_data_adr] * 256 + Once_Data[once_data_adr-1])
                        print(once_data_adr, Once_Data[once_data_adr-1], Once_Data[once_data_adr])
                        ram_koushu = Once_Data[once_data_adr] * 256 + Once_Data[once_data_adr - 1]
                    else:
                        ram_koushu = Once_Data[once_data_adr] * 256 + Once_Data[once_data_adr-1]
                    once_data_adr -= 1
                    # print("第 ", Koushu[0]-1, "口抽吸时间=", once_data_adr)
                    Suction_Duration[Koushu[0]] = (once_data_adr - 1)*0.4
                    # print("第 ", Koushu[0] - 1, "口抽吸时间=", Suction_Duration[Koushu[0]-1])
                    for ram_adr in range(1, once_data_adr, 1):
                        if Once_Data[ram_adr] != 240:
                            if Once_Data[ram_adr] >= 230:   # 若斜率为负数
                                Once_Data[ram_adr] = (255 - Once_Data[ram_adr]) / 10000
                            else:
                                Once_Data[ram_adr] = Once_Data[ram_adr] / 10000
                            Sheet0_data.cell(ram_num-(once_data_adr+2)+ram_adr, 2).value = Once_Data[ram_adr]
                        if Once_Data[ram_adr-1] != 250:
                            if Once_Data[ram_adr] >= DRY_BURNING_THTRSHOLD_OTHER and Once_Data[ram_adr] != 240:
                                if Once_Data[ram_adr+1] == 240:
                                    K_write_En = 1
                                else:
                                    Error_flag = 0
                        else:
                            if Once_Data[ram_adr] >= DRY_BURNING_THTRSHOLD_FIRST and Once_Data[ram_adr] != 240:
                                if Once_Data[ram_adr+1] == 240:
                                    K_write_En = 1
                                else:
                                    Error_flag = 2
                        if Error_flag != 0:
                            Error_flag = 0
                            print("斜率存储/判断错误!!!", "口数=", Koushu[0]-1, Once_Data[ram_adr], Once_Data[ram_adr+1])
                        if K_write_En == 1:
                            K_write_En = 0
                            fills = PatternFill("solid", fgColor='FFA500')
                            Sheet0_data.cell(array_3_adr, 3, Koushu[0] - 1).fill = fills        # 写入当前口数
                            array_3_adr += 1
                            fills = PatternFill("solid", fgColor='87CEFA')
                            Sheet0_data.cell(array_3_adr, 3, ram_adr*0.4).fill = fills          # 写入当前口数
                            array_3_adr += 1
                            fills = PatternFill("solid", fgColor='8FBC8F')
                            Sheet0_data.cell(array_3_adr, 3, Once_Data[ram_adr]).fill = fills    # 写入斜率K
                            array_3_adr += 1
                    Sheet0_data.cell(ram_num-2, 2).value = Suction_Duration[Koushu[0]-1]
                Once_Data[once_data_adr] = rom_data     # 将此次抽吸的口数的起始帧写入
                once_data_adr = 1
            else:
                Once_Data[once_data_adr] = rom_data
                once_data_adr += 1
            if Koushu[0] == 0:      # 防止第一次读取存储数据时出现数据为0时，导致 Once_Data[]存储的数据异常
                once_data_adr = 0
        else:
            if isinstance(rom_data, int):       # isinstance(a,(str,int,list)) 是元组中的一个返回 True
                Once_Data[once_data_adr] = rom_data
                once_data_adr += 1
    if once_data_adr != 1:
        once_data_adr -= 1
        if ((Once_Data[once_data_adr] * 256 + Once_Data[once_data_adr - 1]) - 1) != ram_koushu and ram_koushu != 0:
            print("口数存储错误!!!", "口数=", Koushu[0] - 1, "记录上一口数", ram_koushu, "当前口数",
                  Once_Data[once_data_adr] * 256 + Once_Data[once_data_adr - 1])
            print(once_data_adr, Once_Data[once_data_adr - 1], Once_Data[once_data_adr])
            ram_koushu = Once_Data[once_data_adr] * 256 + Once_Data[once_data_adr - 1]
        else:
            ram_koushu = Once_Data[once_data_adr] * 256 + Once_Data[once_data_adr - 1]
        once_data_adr -= 1
        Suction_Duration[Koushu[0]+1] = (once_data_adr - 1) * 0.4
        for ram_adr in range(1, once_data_adr, 1):
            if Once_Data[ram_adr] != 240:
                if Once_Data[ram_adr] >= 230:  # 若斜率为负数
                    Once_Data[ram_adr] = (255 - Once_Data[ram_adr]) / 10000
                else:
                    Once_Data[ram_adr] = Once_Data[ram_adr] / 10000
                Sheet0_data.cell(ram_num - (once_data_adr + 2) + ram_adr, 2).value = Once_Data[ram_adr]
            if Once_Data[ram_adr - 1] != 250:
                if Once_Data[ram_adr] >= DRY_BURNING_THTRSHOLD_OTHER and Once_Data[ram_adr] != 240:
                    if Once_Data[ram_adr + 1] == 240:
                        K_write_En = 1
                    else:
                        Error_flag = 0
            else:
                if Once_Data[ram_adr] >= DRY_BURNING_THTRSHOLD_FIRST and Once_Data[ram_adr] != 240:
                    if Once_Data[ram_adr + 1] == 240:
                        K_write_En = 1
                    else:
                        Error_flag = 2
            if Error_flag != 0:
                Error_flag = 0
                print("斜率存储/判断错误!!!", "口数=", Koushu[0], Once_Data[ram_adr], Once_Data[ram_adr + 1])
            if K_write_En == 1:
                K_write_En = 0
                fills = PatternFill("solid", fgColor='FFA500')
                Sheet0_data.cell(array_3_adr, 3, Koushu[0]).fill = fills  # 写入当前口数
                array_3_adr += 1
                fills = PatternFill("solid", fgColor='87CEFA')
                Sheet0_data.cell(array_3_adr, 3, ram_adr * 0.4).fill = fills  # 写入当前口数
                array_3_adr += 1
                fills = PatternFill("solid", fgColor='8FBC8F')
                Sheet0_data.cell(array_3_adr, 3, Once_Data[ram_adr]).fill = fills  # 写入斜率K
                array_3_adr += 1
        Sheet0_data.cell(ram_num - 2, 2).value = Suction_Duration[Koushu[0]+1]
        # print("最后一口-》", Suction_Duration[Koushu[0] - 1])


if __name__ == '__main__':
    for num in range(0, xlsx_sheet_number, 1):   # xlsx_sheet_number
        Sheet0_data = Rd[Rd.sheetnames[num]]  # 获取工作表0内的所有数据 创建数组变量的目的是防止后面使用该变量时，编译器判定为创建新变量
        if Sheet0_data.max_row != 1:
            Koushu[0] = 0
            xlsx_read()
            for num1 in range(0, Koushu[0]+2, 1):
                if (Suction_Duration[num1] >= 0.4) and (Suction_Duration[num1] <= 0.5):
                    Save_Koushu_Time_Statistics[0] += 1
                elif Suction_Duration[num1] >= 0.8 and (Suction_Duration[num1] <= 0.9):
                    Save_Koushu_Time_Statistics[1] += 1
                elif Suction_Duration[num1] >= 1.2 and (Suction_Duration[num1] <= 1.3):
                    Save_Koushu_Time_Statistics[2] += 1
                elif Suction_Duration[num1] >= 1.6 and (Suction_Duration[num1] <= 1.7):
                    Save_Koushu_Time_Statistics[3] += 1
                elif Suction_Duration[num1] >= 2.0 and (Suction_Duration[num1] <= 2.1):
                    Save_Koushu_Time_Statistics[4] += 1
                Suction_Duration[num1] = 0
            Sheet0_data.cell(1, 4, Save_Koushu_Time_Statistics[0])
            Sheet0_data.cell(2, 4, Save_Koushu_Time_Statistics[1])
            Sheet0_data.cell(3, 4, Save_Koushu_Time_Statistics[2])
            Sheet0_data.cell(4, 4, Save_Koushu_Time_Statistics[3])
            Sheet0_data.cell(5, 4, Save_Koushu_Time_Statistics[4])
            # print(Save_Koushu_Time_Statistics[0])
            # print(Save_Koushu_Time_Statistics[1])
            # print(Save_Koushu_Time_Statistics[2])
            # print(Save_Koushu_Time_Statistics[3])
            # print(Save_Koushu_Time_Statistics[4])
            Sheet0_data.cell(6, 4, "累计时长：")
            Sheet0_data.cell(6, 5, Save_Koushu_Time_Statistics[0]*0.4+Save_Koushu_Time_Statistics[1]*0.8+Save_Koushu_Time_Statistics[2]*1.2+Save_Koushu_Time_Statistics[3]*1.6+Save_Koushu_Time_Statistics[4]*2.0)
            Save_Koushu_Time_Statistics[0] = 0
            Save_Koushu_Time_Statistics[1] = 0
            Save_Koushu_Time_Statistics[2] = 0
            Save_Koushu_Time_Statistics[3] = 0
            Save_Koushu_Time_Statistics[4] = 0
            print(Rd.sheetnames[num], "->: ", Koushu[0])

    Rd.save(filename)
    print("end")
