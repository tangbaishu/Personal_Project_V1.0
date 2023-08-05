# import xlrd  # 引入Excel库的xlrd
# import numpy
# import matplotlib
# import scipy
# import xlutils.copy
import openpyxl
# import sys
#
# import matplotlib.pyplot as plt
# import numpy as np
import TLB

# 软件功能说明：
#    读取Excel表格内的发热丝阻值数据，基于该数据，在对应时间段内对阻值取平均值，并判断是否超出干烧阈值
#

Cycle_Mean = 0          # 间隔 Mean_Time ms周期取平均
if Cycle_Mean == 1:     # 对应持续打印2s内的数据
    Mean_Time = 200  # 对xms内的数据取平均值
else:   # 对应周期打印200~400的数据
    Mean_Time = 200  # 对xms内的数据取平均值
K_Threshold_Value = 0.003   # 触发干烧阈值


filename = r'D:\Desktop\Python\RAM_Excel_Data.xlsx'
# wb = xlutils.copy.copy(Excel_Data)
# ws = wb.get_sheet(0)
Rd = openpyxl.load_workbook(filename)      # 加载存在的excel文件，与Workbook()的区别在于，load_workbook（）未一个方法，一般是加载已存在的文件，Workbook()是创建一个类对象，用于创建一个空白的excel（）
# print(Rd.get_sheet_names())
Sheet0_data = Rd['Sheet5']          # 获取工作表0内的所有数据
Array_Len = Sheet0_data.max_row+1     # 获取该工作表内的最大行
List_Len = Sheet0_data.max_column+1   # 获取该工作表内的最大列
print("Array_Len=", Array_Len)
print("List_Len=", List_Len)
Smoke_R_Value = [0]*Array_Len       # 暂存excel读取的阻值
Smoke_Time = [0]*Array_Len          # 暂存excel读取的阻值获取时间
Mean_R_Value = [0]*Array_Len        # 暂存200ms内的平均值
K_Value = [0.0]*Array_Len             # 暂存200ms内的斜率值
K_Value_Max = [0.0]*Array_Len         # 暂存斜率最大值
Cycler_Data = [0]*Array_Len         # 暂存斜率读取周期
List_Adr = [1]                      # 暂存当前正在写入/读取的列地址
Valid_Data_len = [0]*4              # adr[0]:Smoke_R_Value  adr[1]:Smoke_Time   adr[2]:Mean_R_Value     adr[3]:K_Value
Check_Data = [0]*2


# 对应读取200ms~400ms数据值
def Read_Data_Way0(list_adr):
    check = 0
    rom_data = 0
    rom_ASCII = 0
    for num in range(1, Array_Len, 1):
        rom_data = Sheet0_data.cell(num, list_adr).value   # 提取Sheets0内第1列num行的数据  注：该库数据读取从1开始。因此第一列对应列表A
        rom_data = str(rom_data)    # 将获取的数据转换成字符
        if rom_data == 'None':      # 判断读取的数据是否为空数据
            print("Read_None->", num)
        else:
            if check != 3:
                check = 0
            for adr in range(0, len(rom_data), 1):  # len(rom_data)获取字符长度
                rom_ASCII = ord(rom_data[adr])  # 将字符转换成ASCII码
                if rom_ASCII == 61:
                    check = 1
                    if ord(rom_data[adr-1]) == 82:  # R
                        check = 2
                    if ord(rom_data[adr+1]) == 45:
                        check = 3
            if check == 2:  # R=
                Smoke_R_Value[Valid_Data_len[0]] = TLB.string_open_sql(rom_data)  # 提取字符串内的数字
                if (Smoke_R_Value[Valid_Data_len[0]] >= 0) and (Smoke_R_Value[Valid_Data_len[0]] < 3):
                    Valid_Data_len[0] += 1
                    Check_Data[0] = 1
                else:
                    Smoke_R_Value[Valid_Data_len[0]] = 0
                # print("Smoke_R_Value[", Valid_Data_len[0]-1, "] = ", Smoke_R_Value[Valid_Data_len[0]-1], "->", num)
            elif check == 0:
                if (ord(rom_data[0]) >= 48) and (ord(rom_data[0]) <= 57):     # 提取数字
                    Smoke_Time[Valid_Data_len[1]] = TLB.string_open_sql(rom_data)  # 提取字符串内的数字
                    if Smoke_Time[Valid_Data_len[1]] >= 0:
                        Valid_Data_len[1] += 1
                        Check_Data[0] = 1
                    else:
                        Smoke_Time[Valid_Data_len[1]] = 0
                # print("Smoke_Time[", Valid_Data_len[1]-1, "] = ", Smoke_Time[Valid_Data_len[1]-1], "->", num)


# 对应持续打印2s内的数据
def Read_Data_Way1(list_adr):
    check = 0
    rom_data = 0
    rom_ASCII = 0
    for num in range(1, Array_Len, 1):
        rom_data = Sheet0_data.cell(num, list_adr).value  # 提取Sheets0内第1列num行的数据  注：该库数据读取从1开始。因此第一列对应列表A
        rom_data = str(rom_data)  # 将获取的数据转换成字符
        check = 3
        if rom_data == 'None':  # 判断读取的数据是否为空数据
            print("Read_None->", num)
            return
        else:
            for adr in range(0, len(rom_data), 1):  # len(rom_data)获取字符长度
                rom_ASCII = ord(rom_data[adr])  # 将字符转换成ASCII码
                if rom_ASCII == 61:             # =
                    check = 1
                    if ord(rom_data[adr - 1]) == 101:   # e
                        check = 2
                    if ord(rom_data[adr - 1]) == 115:   # s
                        check = 0
            if check == 2:  # e=
                Smoke_R_Value[Valid_Data_len[0]] = TLB.string_open_sql(rom_data)  # 提取字符串内的数字
                if Smoke_R_Value[Valid_Data_len[0]] >= 0:
                    Valid_Data_len[0] += 1
                    Check_Data[0] = 1
                else:
                    Smoke_R_Value[Valid_Data_len[0]] = 0
                # print("Smoke_R_Value[", Valid_Data_len[0]-1, "] = ", Smoke_R_Value[Valid_Data_len[0]-1], "->", num)
            elif check == 0:    # s=
                Smoke_Time[Valid_Data_len[1]] = TLB.string_open_sql(rom_data)  # 提取字符串内的数字
                if Smoke_Time[Valid_Data_len[1]] >= 0:
                    Valid_Data_len[1] += 1
                    Check_Data[0] = 1
                else:
                    Smoke_Time[Valid_Data_len[1]] = 0
                # print("Smoke_Time[", Valid_Data_len[1]-1, "] = ", Smoke_Time[Valid_Data_len[1]-1], "->", num)


# 200~400ms时间周期数据处理
def Time_2to4ms_Data_Dispose():
    time_200ms = 0
    base = 0    # 基数
    for num in range(0, Valid_Data_len[1]+1, 1):              # 获取平均数
        if Smoke_Time[num] <= Mean_Time:
            if time_200ms == 1:
                if base != 0:
                    Mean_R_Value[Valid_Data_len[2]] /= base
                    Valid_Data_len[2] += 1
                    base = 0
                time_200ms = 0
            Mean_R_Value[Valid_Data_len[2]] += Smoke_R_Value[num]
            base += 1
        elif Smoke_Time[num] <= Mean_Time*2:
            if time_200ms == 0:
                if base != 0:
                    Mean_R_Value[Valid_Data_len[2]] /= base
                    Valid_Data_len[2] += 1
                    base = 0
                time_200ms = 1
            Mean_R_Value[Valid_Data_len[2]] += Smoke_R_Value[num]
            base += 1
        if num == Valid_Data_len[1]:
            Mean_R_Value[Valid_Data_len[2]] /= base
            Valid_Data_len[2] += 1
            break
        # print(Smoke_Time[num], Smoke_R_Value[num], Mean_R_Value[Valid_Data_len[2]], base)
    for num in range(0, Valid_Data_len[2], 2):              # 获取斜率K
        K_Value[Valid_Data_len[3]] = (Mean_R_Value[num+1] - Mean_R_Value[num])/2
        if K_Value[Valid_Data_len[3]] >= 1:
            K_Value[Valid_Data_len[3]] = 0
        else:
            Valid_Data_len[3] += 1
    for num in range(0, Valid_Data_len[3], 1):              # 获取最大斜率K，以及触发阈值后的斜率
        if K_Value_Max[0] < K_Value[num]:
            K_Value_Max[0] = K_Value[num]
        if K_Value[num] > K_Threshold_Value:
            K_Value_Max[num] = K_Value[num]
    print("K_Value_Max = ", K_Value_Max[0])


# 每200ms获取平均值
def Time_200ms_mean():
    time_200ms = 0
    base = 0  # 基数
    for num in range(0, Valid_Data_len[1]+1, 1):  # 获取平均数
        if Smoke_Time[num] <= time_200ms+Mean_Time:
            Mean_R_Value[Valid_Data_len[2]] += Smoke_R_Value[num]
            base += 1
        else:
            Mean_R_Value[Valid_Data_len[2]] /= base
            time_200ms += Mean_Time
            Valid_Data_len[2] += 1
            base = 1
            Mean_R_Value[Valid_Data_len[2]] += Smoke_R_Value[num]
        if num == Valid_Data_len[1]:
            if Smoke_Time[num] == 0:
                Mean_R_Value[Valid_Data_len[2]] -= Smoke_R_Value[num]
                if base != 1:
                    base -= 1
            Mean_R_Value[Valid_Data_len[2]] /= base
            # print(Smoke_Time[num], Smoke_R_Value[num], Mean_R_Value[Valid_Data_len[2]], base)
            Valid_Data_len[2] += 1
            break
        # print(Smoke_Time[num], Smoke_R_Value[num], Mean_R_Value[Valid_Data_len[2]], base)
    for num in range(0, Valid_Data_len[2], 2):  # 获取斜率K
        K_Value[Valid_Data_len[3]] = (Mean_R_Value[num + 1] - Mean_R_Value[num]) / 2
        if K_Value[Valid_Data_len[3]] >= 1:
            K_Value[Valid_Data_len[3]] = 0
        else:
            Valid_Data_len[3] += 1
    for num in range(0, Valid_Data_len[3], 1):  # 获取最大斜率K，以及触发阈值后的斜率
        if K_Value_Max[0] < K_Value[num]:
            K_Value_Max[0] = K_Value[num]
        if K_Value[num] > K_Threshold_Value:
            K_Value_Max[num] = K_Value[num]
    # print("K_Value_Max = ", K_Value_Max[0])


def Write_Data():
    if Check_Data[1] == 0:
        List_Adr[0] += 1
        Sheet0_data.insert_cols(List_Adr[0], 5)                                 # 插入x行
    # Sheet0_data.row_dimensions[1].height = 100                                # 设置行高
    Sheet0_data.column_dimensions['A'].width = 22                               # 设置列宽 测试发现该列宽并非设置A列会作用在该工作表全局内，此时设置其它列可以修改该列的列宽，但是未明确设置的会默认为A列的值
    for num in range(0, len(Valid_Data_len), 1):
        Valid_Data_len[num] += 1
    if List_Adr[0] == 2:
        Sheet0_data.cell(1, List_Adr[0]).value = "Winsemi"
        for num in range(2, Valid_Data_len[1]+1, 1):
            Sheet0_data.cell(num, List_Adr[0]).value = Smoke_Time[num-2]            # 将时间数据写入x列
            # print("Smoke_Time[", num-1, "]=", Smoke_Time[num-1])
            Smoke_Time[num - 2] = 0
    else:
        for num in range(1, Valid_Data_len[1], 1):
            Sheet0_data.cell(num, List_Adr[0]).value = Smoke_Time[num-1]            # 将时间数据写入x列
            # print("Smoke_Time[", num-1, "]=", Smoke_Time[num-1])
            Smoke_Time[num - 1] = 0
    List_Adr[0] += 1
    for num in range(1, Valid_Data_len[0], 1):
        Sheet0_data.cell(num, List_Adr[0]).value = Smoke_R_Value[num-1]         # 将阻值数据写入x列
        # print("Smoke_R_Value[", num - 1, "]=", Smoke_R_Value[num - 1])
        Smoke_R_Value[num - 1] = 0
    List_Adr[0] += 1
    for num in range(1, Valid_Data_len[2], 1):
        Sheet0_data.cell(num, List_Adr[0]).value = Mean_R_Value[num-1]          # 将平均阻值数据写入x列
        # print("Mean_R_Value[", num-1, "]=", Mean_R_Value[num-1])
        Mean_R_Value[num - 1] = 0
    List_Adr[0] += 1
    for num in range(1, Valid_Data_len[3], 1):
        Sheet0_data.cell(num, List_Adr[0]).value = K_Value[num - 1]             # 将阻值变化斜率K写入x列
        # print("K_Value[", num-1, "]=", K_Value[num-1])
        K_Value[num - 1] = 0
    List_Adr[0] += 1
    for num in range(1, Valid_Data_len[3], 1):
        Sheet0_data.cell(num, List_Adr[0]).value = K_Value_Max[num-1]           # 将最大斜率K，以及触发阈值后的斜率数据写入x列
        # print("K_Value_Max[", num-1, "]=", K_Value_Max[num-1])
        K_Value_Max[num - 1] = 0


if __name__ == '__main__':
    if Cycle_Mean == 1:  # 对应持续打印2s内的数据
        print("Time_200ms_mean")
        # Check_str = Sheet0_data.cell(1, 2).value  # 提取Sheets0内第1列num行的数据  注：该库数据读取从1开始。因此第一列对应列表A
        # Check_str = str(Check_str)  # 将获取的数据转换成字符
        # if Check_str == "Winsemi":
        #     Check_Data[1] = 1  # 基于第二列数据判断之前是否以执行过数据处理，若执行过数据则 Check_Data[0] = 1，否则为0
        for new_list_adr in range(List_Adr[0], List_Len, 1):    # List_Len
            print("List_Adr[0]->", List_Adr[0])
            Check_Data[0] = 0
            Read_Data_Way1(List_Adr[0])
            if Check_Data[0] == 1:
                Time_200ms_mean()    # 对xms内的数据取平均值
                Write_Data()
                List_Adr[0] += 1
                Valid_Data_len[0] = 0
                Valid_Data_len[1] = 0
                Valid_Data_len[2] = 0
                Valid_Data_len[3] = 0
            else:
                print("无效数据")
                List_Adr[0] += 1
    else:  # 对应周期打印200~400的数据
        print("Time_2to4ms_Data_Dispose")
        Read_Data_Way0(List_Adr[0])
        Time_2to4ms_Data_Dispose()
        Write_Data()
    Rd.save(filename)

    print("end")
