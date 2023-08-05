# import numpy
# import matplotlib  # Python 绘图库
# import scipy

import openpyxl  # Python 下的Excel库

import matplotlib.pyplot as plt     # 导入 matplotlib 库中的 pyplot模块 并重命名为 plt
import numpy as np  # 内含维度数组、矩阵运算、数学函数库
# import xlsxwriter as xw

import TLB
# import DataBase

Array_Len = 1234

filename = r"D:\Desktop\WorkFile\E-Cigarette_Data Model\Discharge Characteristic Data\RAM_DATA.xlsx"    # 将需要读取的文件地址 放置该变量
Excel_Data = openpyxl.load_workbook(filename)  # 加载对应文件，并将数据存入变量
Excel_Sheets0_Data = Excel_Data['Sheet1']   # 获取对应工作表数据

num = 0
excel_char_data = ['A'] * Array_Len
excel_data0 = [0] * Array_Len
excel_data1 = [0] * Array_Len

for num in range(1, Array_Len, 1):
    if num == 1:
        excel_char_data[num] = str(Excel_Sheets0_Data.cell(1, 1).value)  # 提取Sheets0内第1行1列的数据  注：该库数据读取从1开始。因此第一列对应列表 A1
        excel_data0[num] = TLB.string_open_sql(excel_char_data[num])
        num += 1

    excel_char_data[num] = str(Excel_Sheets0_Data.cell(num, 1).value)  # 提取Sheets0内第num行A列的数据  注：该库数据读取从1开始。因此第一列对应列表 Anum
    # print(excel_char_data[num])
    excel_data0[num] = TLB.string_open_sql(excel_char_data[num])
    # print(excel_data0[num])

for num in range(0, Array_Len, 1):
    if num == 0:
        excel_char_data[num] = str(Excel_Sheets0_Data.cell(1, 2).value)     # 提取Sheets0内第1行2列的数据  注：该库数据读取从1开始。因此第一列对应列表 B1
        excel_data1[num] = TLB.string_open_sql(excel_char_data[num])
        num += 1
    # if Excel_Sheets0_Data[num][1] == '\0':
    #     print("num=", num)
    #     break
    excel_char_data[num] = str(Excel_Sheets0_Data.cell(num, 2).value)        # 提取Sheets0内第num行2列的数据  注：该库数据读取从1开始。因此第一列对应列表 Bnum
    # print(excel_char_data[num])
    excel_data1[num] = TLB.string_open_sql(excel_char_data[num])
    # print(excel_data1[num])

plt.plot(excel_data1, excel_data0, '.')   # 绘制color 曲线颜色、linestyle曲线样式
# plt.show()

x0_len = 0
x_len = [250, 60, 0]
number_of_fits = [8, 8, 6]
k = [0]*len(x_len)
adr = 0
data_x0 = [0] * Array_Len
data_y0 = [0] * Array_Len


for duty in range(0, len(x_len), 1):
    for num in range(0, Array_Len, 1):
        if duty == 0:
            if excel_data1[num] > x_len[duty]:
                data_x0[num] = excel_data1[num]
                data_y0[num] = excel_data0[num]
            else:
                adr = num
                break
        else:
            if (excel_data1[num + adr] > x_len[duty]) and ((num + adr) < (Array_Len-1)):
                data_x0[num] = excel_data1[num + adr]
                data_y0[num] = excel_data0[num + adr]
            elif excel_data1[num + adr] < x_len[duty]:
                adr = adr + num
                break
            elif (num + adr) == (Array_Len-1):
                data_x0[num] = excel_data1[num + adr]
                data_y0[num] = excel_data0[num + adr]
                num += 1
                break
        # print(data_x0[num], num+adr)  # 显示待拟合曲线的X轴数据
    x0 = [0] * num
    y0 = [0] * num
    length = num
    print("num=", num)
    print("duty=", duty)

    for num in range(0, length, 1):
        x0[num] = data_x0[num]
        y0[num] = data_y0[num]
        # print(x0[num])  # 显示待拟合曲线的X轴数据
    # plt.plot(x0, y0, '-')
    z0 = np.polyfit(x0, y0, number_of_fits[duty])
    k[duty] = np.poly1d(z0)  # 使用次数合成多项式
    print(z0)
    print("y=", k[duty])
    y0 = k[duty](x0)
    # plt.plot(x0, y0, color='r')
    # plt.show()


for num in range(0, Array_Len, 1):
    if excel_data1[num] == 0:
        data_y0[num] = excel_data0[Array_Len-1]
    else:
        if excel_data1[num] > x_len[0]:
            data_y0[num] = k[0](excel_data1[num])
        elif excel_data1[num] > x_len[1]:
            data_y0[num] = k[1](excel_data1[num])
        elif excel_data1[num] > x_len[2]:
            data_y0[num] = k[2](excel_data1[num])
    # elif excel_data1[num] > x_len[3]:
    #     data_y0[num] = k[3](excel_data1[num])
    # ws.write(num, 2, data_y0[num])

plt.plot(excel_data1, data_y0, color='r')

if __name__ == '__main__':
    # Excel_Data.save(filename)
    plt.show()
