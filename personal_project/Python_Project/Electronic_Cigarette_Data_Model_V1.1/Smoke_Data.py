import xlrd  # 引入Excel库的xlrd
import numpy
import matplotlib
import scipy
import xlutils.copy
import openpyxl
import sys

import matplotlib.pyplot as plt
import numpy as np
import TLB

Cycle_Mean = 1          # 间隔 Mean_Time ms周期取平均
if Cycle_Mean == 1:     # 对应持续打印2s内的数据
    Mean_Time = 200  # 对xms内的数据取平均值
else:   # 对应周期打印200~400的数据
    Mean_Time = 200  # 对xms内的数据取平均值


filename = r'C:\Users\WX\Desktop\Python\2S_20230210_V3.0.xlsx'
# wb = xlutils.copy.copy(Excel_Data)
# ws = wb.get_sheet(0)
Rd = openpyxl.load_workbook(filename)      # 加载存在的excel文件，与Workbook()的区别在于，load_workbook（）未一个方法，一般是加载已存在的文件，Workbook()是创建一个类对象，用于创建一个空白的excel（）
# print(Rd.get_sheet_names())
Sheet0_data = Rd['KTR-8W-A0']          # 获取工作表0内的所有数据
Array_Len = Sheet0_data.max_row+1     # 所有获取的数组长度
print(Array_Len)
Smoke_R_Value = [0]*Array_Len       # 暂存excel读取的阻值
Smoke_Time = [0]*Array_Len          # 暂存excel读取的阻值获取时间
Mean_R_Value = [0]*Array_Len        # 暂存200ms内的平均值
K_Value = [0]*Array_Len             # 暂存200ms内的斜率值
K_Value_Max = [0]                   # 暂存斜率最大值
Valid_Data_len = [0]*4              # adr[0]:Smoke_R_Value  adr[1]:Smoke_Time   adr[2]:Mean_R_Value     adr[3]:K_Value


# 对应读取200ms~400ms数据值
def Read_Data_Way0():
    check = 0
    rom_data = 0
    rom_ASCII = 0
    for num in range(1, Array_Len, 1):
        rom_data = Sheet0_data.cell(num, 1).value   # 提取Sheets0内第1列num行的数据  注：该库数据读取从1开始。因此第一列对应列表A
        rom_data = str(rom_data)    # 将获取的数据转换成字符
        if rom_data == 'None':      # 判断读取的数据是否为空数据
            print("Read_None->", num)
        else:
            if check != 3:
                check = 0
            for adr in range(0, len(rom_data), 1):  # len(rom_data)获取字符长度
                rom_ASCII = ord(rom_data[adr])  # 将字符转换成ASCII码
                if rom_ASCII == 61:
                    check = 1
                    if ord(rom_data[adr-1]) == 82:
                        check = 2
                    if ord(rom_data[adr+1]) == 45:
                        check = 3
            if check == 2:
                Smoke_R_Value[Valid_Data_len[0]] = TLB.string_open_sql(rom_data)  # 提取字符串内的数字
                if Smoke_R_Value[Valid_Data_len[0]] >= 0:
                    Valid_Data_len[0] += 1
                else:
                    Smoke_R_Value[Valid_Data_len[0]] = 0
                # print("Smoke_R_Value[", Valid_Data_len[0]-1, "] = ", Smoke_R_Value[Valid_Data_len[0]-1], "->", num)
            elif check == 0:
                if (ord(rom_data[0]) >= 48) and (ord(rom_data[0]) <= 57):     # 提取数字
                    Smoke_Time[Valid_Data_len[1]] = TLB.string_open_sql(rom_data)  # 提取字符串内的数字
                    if Smoke_Time[Valid_Data_len[1]] >= 0:
                        Valid_Data_len[1] += 1
                    else:
                        Smoke_Time[Valid_Data_len[1]] = 0
                # print("Smoke_Time[", Valid_Data_len[1]-1, "] = ", Smoke_Time[Valid_Data_len[1]-1], "->", num)


# 对应持续打印2s内的数据
def Read_Data_Way1():
    check = 0
    rom_data = 0
    rom_ASCII = 0
    for num in range(1, Array_Len, 1):
        rom_data = Sheet0_data.cell(num, 1).value  # 提取Sheets0内第1列num行的数据  注：该库数据读取从1开始。因此第一列对应列表A
        rom_data = str(rom_data)  # 将获取的数据转换成字符
        if rom_data == 'None':  # 判断读取的数据是否为空数据
            print("Read_None->", num)
            return
        else:
            for adr in range(0, len(rom_data), 1):  # len(rom_data)获取字符长度
                rom_ASCII = ord(rom_data[adr])  # 将字符转换成ASCII码
                if rom_ASCII == 61:             # =
                    check = 1
                    if ord(rom_data[adr - 1]) == 101:   # e
                        check = 2
                    if ord(rom_data[adr - 1]) == 115:   # s
                        check = 0
            if check == 2:
                Smoke_R_Value[Valid_Data_len[0]] = TLB.string_open_sql(rom_data)  # 提取字符串内的数字
                if Smoke_R_Value[Valid_Data_len[0]] >= 0:
                    Valid_Data_len[0] += 1
                else:
                    Smoke_R_Value[Valid_Data_len[0]] = 0
                print("Smoke_R_Value[", Valid_Data_len[0]-1, "] = ", Smoke_R_Value[Valid_Data_len[0]-1], "->", num)
            elif check == 0:
                Smoke_Time[Valid_Data_len[1]] = TLB.string_open_sql(rom_data)  # 提取字符串内的数字
                if Smoke_Time[Valid_Data_len[1]] >= 0:
                    Valid_Data_len[1] += 1
                else:
                    Smoke_Time[Valid_Data_len[1]] = 0
                print("Smoke_Time[", Valid_Data_len[1]-1, "] = ", Smoke_Time[Valid_Data_len[1]-1], "->", num)


# 200~400ms时间周期数据处理
def Time_2to4ms_Data_Dispose():
    time_200ms = 0
    base = 0    # 基数
    for num in range(0, Valid_Data_len[1], 1):
        if Smoke_Time[num] <= Mean_Time:
            if time_200ms == 1:
                if base != 0:
                    Mean_R_Value[Valid_Data_len[2]] /= base
                    Valid_Data_len[2] += 1
                    base = 0
                time_200ms = 0
            Mean_R_Value[Valid_Data_len[2]] += Smoke_R_Value[num]
            base += 1
        else:
            if time_200ms == 0:
                if base != 0:
                    Mean_R_Value[Valid_Data_len[2]] /= base
                    Valid_Data_len[2] += 1
                    base = 0
                time_200ms = 1
            Mean_R_Value[Valid_Data_len[2]] += Smoke_R_Value[num]
            base += 1
    for num in range(0, Valid_Data_len[2], 2):
        K_Value[Valid_Data_len[3]] = (Mean_R_Value[num+1] - Mean_R_Value[num])/2
        if K_Value[Valid_Data_len[3]] >= 1:
            K_Value[Valid_Data_len[3]] = 0
        else:
            Valid_Data_len[3] += 1
    for num in range(0, Valid_Data_len[3], 1):
        if K_Value_Max[0] < K_Value[num]:
            K_Value_Max[0] = K_Value[num]
    print("K_Value_Max = ", K_Value_Max[0])


# 每200ms获取平均值
def Time_200ms_mean():
    Sheet0_data.insert_cols(2, 5)


def Write_Data():
    for num in range(0, len(Valid_Data_len), 1):
        Valid_Data_len[num] += 1
    for num in range(1, Valid_Data_len[1], 1):
        Sheet0_data.cell(num, 2).value = Smoke_Time[num-1]          # 将时间数据写入B列
        # print("Smoke_Time[", num-1, "]=", Smoke_Time[num-1])
    for num in range(1, Valid_Data_len[0], 1):
        Sheet0_data.cell(num, 3).value = Smoke_R_Value[num-1]       # 将阻值数据写入C列
        # print("Smoke_R_Value[", num - 1, "]=", Smoke_R_Value[num - 1])
    for num in range(1, Valid_Data_len[2], 1):
        Sheet0_data.cell(num, 4).value = Mean_R_Value[num-1]        # 将阻值数据写入D列
        # print("Mean_R_Value[", num-1, "]=", Mean_R_Value[num-1])
    for num in range(1, Valid_Data_len[3], 1):
        Sheet0_data.cell(num, 5).value = K_Value[num - 1]           # 将阻值数据写入E列
        # print("K_Value[", num-1, "]=", K_Value[num-1])
    Sheet0_data.cell(1, 6).value = K_Value_Max[0]                   # 将阻值数据写入F列


if __name__ == '__main__':
    if Cycle_Mean == 1:  # 对应持续打印2s内的数据
        Read_Data_Way1()  # 对xms内的数据取平均值
    else:  # 对应周期打印200~400的数据
        Read_Data_Way0()
    Time_2to4ms_Data_Dispose()
    Time_200ms_mean()
    Write_Data()
    # Rd.save(filename)

    print("end")
